from flask import Flask, render_template, Response
import cv2
from datetime import datetime
import csv
from model.test_fer2013 import detect_emotion
import os
from datetime import datetime

app = Flask(__name__)
camera = cv2.VideoCapture(0)

# Khởi tạo nhãn cảm xúc trước đó
last_emotion = ""

def gen_frames():
    global last_emotion
    os.makedirs("static", exist_ok=True)

    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            # ✅ Dự đoán 1 lần duy nhất
            processed_frame, emotion = detect_emotion(frame.copy())

            # ✅ Chỉ lưu nếu cảm xúc thay đổi
            if emotion and emotion != last_emotion:
                last_emotion = emotion

                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename_only = f"history_{timestamp}.jpg"
                filepath = os.path.join("static", filename_only)

                # ✅ Lưu ảnh đã có cảm xúc vẽ lên (processed_frame)
                cv2.imwrite(filepath, processed_frame)

                # ✅ Ghi đúng nhãn đã dự đoán
                with open('history.csv', 'a', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow([timestamp, emotion, filename_only])

            # ✅ Stream ảnh đã được vẽ cảm xúc
            _, buffer = cv2.imencode('.jpg', processed_frame)
            frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/history')
def history():
    history_data = []
    if os.path.exists("history.csv"):
        with open("history.csv", "r") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) == 3:
                    timestamp_raw, emotion, img_path = row
                    try:
                        # Chuyển string -> datetime
                        dt = datetime.strptime(timestamp_raw, "%Y%m%d_%H%M%S")
                        history_data.append((dt, emotion, img_path))
                    except:
                        continue
    history_data.reverse()
    return render_template("history.html", history=history_data)

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

@app.route('/export/excel')
def export_excel():
    if not os.path.exists("history.csv"):
        return "Chưa có dữ liệu để xuất.", 404

    # Tạo workbook và sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Emotion History"

    # Tiêu đề
    headers = ["Ngày", "Giờ", "Cảm xúc", "Ảnh"]
    ws.append(headers)

    # Đọc và xử lý từng dòng lịch sử
    with open("history.csv", "r") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) != 3:
                continue

            timestamp, emotion, img_name = row
            try:
                dt = datetime.strptime(timestamp, "%Y%m%d_%H%M%S")
                date_str = dt.strftime("%d/%m/%Y")
                time_str = dt.strftime("%H:%M:%S")
            except:
                continue

            # Thêm dữ liệu text
            ws.append([date_str, time_str, emotion, ""])

            # Thêm ảnh
            img_path = os.path.join("static", img_name)
            if os.path.exists(img_path):
                img = XLImage(img_path)
                img.width = 120
                img.height = 90
                row_idx = ws.max_row
                ws.row_dimensions[row_idx].height = 70
                ws.add_image(img, f"D{row_idx}")

    # Căn chỉnh cột
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25

    # Lưu file
    file_path = "static/history_export_with_images.xlsx"
    wb.save(file_path)

    return f'<p>✅ Đã xuất Excel kèm ảnh: <a href="/static/history_export_with_images.xlsx" download>📥 Tải tại đây</a></p>'

@app.route('/hello')
def hello():
    return "Trang hoạt động oke!"

# ✅ Đặt sau tất cả route
if __name__ == "__main__":
    app.run(debug=True)
